import openpyxl as xl

def Student_xl(filename):
    wb=xl.load_workbook(filename)
    sheet=wb("Sheet1")
    sum=0
    for j in range(2,sheet.max_row):
        r=1
        for i in range(2,sheet.max_column):
            m=int(input(f"Enter mark{i-1} for {sheet[f'A{r+1}']} : "))
            cell=sheet.cell(j,i)
            cell.value=m
            sum+=m
        r+=1
        cell_avg=sheet.cell(j,5)
        cell_avg.value=sum/3


