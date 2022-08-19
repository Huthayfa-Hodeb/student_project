import openpyxl as xl

def Student_xl(filename):
    max_avr=0
    sum1=0
    r=1
    wb=xl.load_workbook(filename)
    sheet=wb["Sheet1"]
    for j in range(2,sheet.max_row+1):
        if sheet[f"A{j}"].value== None:
            break
        for i in range(2,sheet.max_column-3):
            m=int(input(f"Enter mark {i-1} for {sheet[f'A{r+1}'].value} : "))
            cell=sheet.cell(j,i)
            cell.value=m
            sum1+=m
        r+=1
        cell_avg=sheet.cell(j,5)
        cell_avg.value=sum1/3
        sum1=0
        if cell_avg.value > max_avr:
            max_avr=cell_avg.value
    cell_max_avr=sheet.cell(5,8)
    cell_max_avr.value=max_avr
    wb.save("classgrades.xlsx")